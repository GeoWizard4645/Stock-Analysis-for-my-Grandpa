#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 STOCK ANALYSIS  --  daily watchlist workbook generator
================================================================================

 What this does
 --------------
 Every trading day this script rebuilds, automatically, the exact daily
 checklist that used to be typed by hand off the investing.com / TradingView
 charts -- and then adds a large amount of extra analysis on top.

 The hand-built sheet asked nine questions of every ticker:

     %change            the day's move, straight off the watchlist
     Prev Strength      rank of that move inside its group (1 = strongest)
     Price > Pre-Day    did it close above yesterday's close?
     Volume > Pre-Day   did it trade more shares than yesterday?
     MACD Green > Red   is the MACD line above its signal line?
     Price > 8 EMA      above the fast exponential average?
     Price > 21 EMA     above the slow exponential average?
     Price > 50 DMA     above the 50-day simple average?
     Price > Upper BB   above the top Bollinger Band?
     Price > Lower BB   above the bottom Bollinger Band?
     Price > Mid BB     above the middle Bollinger Band?

 Those columns are reproduced here character-for-character, in the same
 positions (A through M), so the workbook still reads the way it always has.
 Everything from column N rightwards is new: a bull score, a plain-English
 verdict, full OHLCV, RSI, Stochastic, ATR, every moving average, the raw
 Bollinger numbers, %B, band width, 52-week range, multi-period returns, and
 the three top news headlines for each ticker as clickable links.

 The workbook also gains four extra tabs: a Dashboard, a Sector Rotation
 heat map, a Signals tab that lists only the things that changed today, and a
 plain-English guide to every column.

 Indicator settings match the chart layout exactly:
     EMA 8, 21     SMA (DMA) 50, 100, 200     Bollinger 20, 2
     MACD 12, 26, 9              Volume MA 20

 Data source
 -----------
 Yahoo Finance, through the free `yfinance` package. No account, no API key,
 no subscription. Headlines come from Yahoo Finance's public RSS feeds.

 How to run it
 -------------
     pip install yfinance pandas openpyxl requests
     python stock_analysis.py

 That writes  Stock_Analysis_MM-DD.xlsx  into the current folder.

 Useful switches
 ---------------
     python stock_analysis.py --history 10      last 10 sessions as tabs
     python stock_analysis.py --no-news         skip headline fetching (faster)
     python stock_analysis.py --out report.xlsx choose the filename
     python stock_analysis.py --json data.json  also emit JSON (for the website)
     python stock_analysis.py --watchlist w.json  use a different watchlist

================================================================================
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import os
import re
import sys
import warnings
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor
from typing import Any, Dict, List, Optional, Sequence

warnings.filterwarnings("ignore")

try:
    import pandas as pd
    import numpy as np
    import yfinance as yf
except ImportError as exc:  # pragma: no cover
    sys.exit(
        f"Missing a required package ({exc.name}).\n"
        "Run:  pip install yfinance pandas openpyxl requests"
    )

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ==============================================================================
#  DEFAULT WATCHLIST
# ==============================================================================
# Edit here, or keep a watchlist.json next to the script and it wins.

DEFAULT_WATCHLIST: Dict[str, Any] = {
    "groups": [
        {
            "key": "sectors",
            "title": "SECTOR SPDRs",
            "numbered": True,
            "rank": True,
            "tickers": ["XLI", "XLV", "XLF", "XLE", "XLP", "XLY",
                        "XLC", "XLB", "XLK", "XLRE", "XLU"],
        },
        {
            "key": "indices",
            "title": "INDEX / BROAD MARKET",
            "numbered": False,
            "rank": True,
            "tickers": ["SPY", "QQQ", "IWM", "SOXL"],
        },
        {
            "key": "stocks",
            "title": "WATCHLIST",
            "numbered": False,
            "rank": True,
            "tickers": ["AAPL", "AMD", "AMZN", "ARKK", "ARM", "ASTS", "AVGO",
                        "CCJ", "CRWV", "DELL", "GOOG", "IREN", "INTC", "MSFT",
                        "MSTR", "MU", "NBIS", "NFLX", "NOK", "NVDA", "PLTR",
                        "QCOM", "RKLB", "SNDK", "TSLA"],
        },
    ],
    "settings": {
        "ema_fast": 8,
        "ema_slow": 21,
        "sma_periods": [50, 100, 200],
        "bb_period": 20,
        "bb_stdev": 2,
        "macd": [12, 26, 9],
        "volume_ma": 20,
        "rsi_period": 14,
        "atr_period": 14,
        "stoch_period": 14,
        "same_tolerance_pct": 0.10,
        "news_per_ticker": 3,
    },
}

CHART_URL = "https://www.investing.com/charts/stocks-charts"


# ==============================================================================
#  COLOUR PALETTE
# ==============================================================================

INK          = "0B1220"   # near-black navy, the header bar
PANEL        = "16243A"   # group-band navy
SUBTLE       = "F1F5F9"   # zebra stripe
BORDER_GREY  = "CBD5E1"
LABEL_GREY   = "64748B"

GREEN_BG, GREEN_FG = "D9F5E3", "12703A"
RED_BG,   RED_FG   = "FBDDDD", "A81E1E"
AMBER_BG, AMBER_FG = "FFF1CC", "8A5A00"
BLUE_BG,  BLUE_FG  = "DDEBFB", "14507E"
ACCENT             = "0EA5E9"

SECTION_FILLS = {
    "sectors": "E8EEF7",
    "indices": "E9F3EC",
    "stocks":  "F5F0FA",
}


# ==============================================================================
#  INDICATOR MATHS
# ==============================================================================

def ema(series: pd.Series, span: int) -> pd.Series:
    return series.ewm(span=span, adjust=False).mean()


def sma(series: pd.Series, window: int) -> pd.Series:
    return series.rolling(window, min_periods=window).mean()


def rsi(close: pd.Series, period: int = 14) -> pd.Series:
    """Wilder's RSI -- the same one investing.com draws."""
    delta = close.diff()
    gain = delta.clip(lower=0.0)
    loss = (-delta).clip(lower=0.0)
    avg_gain = gain.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0.0, np.nan)
    out = 100.0 - (100.0 / (1.0 + rs))
    return out.fillna(100.0).where(avg_loss.notna(), np.nan)


def atr(high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14) -> pd.Series:
    """Wilder's Average True Range."""
    prev_close = close.shift(1)
    tr = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low - prev_close).abs(),
    ], axis=1).max(axis=1)
    return tr.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()


def stochastic_k(high: pd.Series, low: pd.Series, close: pd.Series,
                 period: int = 14) -> pd.Series:
    lowest = low.rolling(period, min_periods=period).min()
    highest = high.rolling(period, min_periods=period).max()
    span = (highest - lowest).replace(0.0, np.nan)
    return ((close - lowest) / span) * 100.0


def adx(high: pd.Series, low: pd.Series, close: pd.Series, period: int = 14) -> pd.Series:
    """Average Directional Index -- how strong the trend is, regardless of direction."""
    up = high.diff()
    down = -low.diff()
    plus_dm = np.where((up > down) & (up > 0), up, 0.0)
    minus_dm = np.where((down > up) & (down > 0), down, 0.0)
    tr = pd.concat([
        high - low,
        (high - close.shift(1)).abs(),
        (low - close.shift(1)).abs(),
    ], axis=1).max(axis=1)
    atr_ = tr.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    plus_di = 100 * pd.Series(plus_dm, index=high.index).ewm(
        alpha=1 / period, adjust=False, min_periods=period).mean() / atr_
    minus_di = 100 * pd.Series(minus_dm, index=high.index).ewm(
        alpha=1 / period, adjust=False, min_periods=period).mean() / atr_
    dx = ((plus_di - minus_di).abs() / (plus_di + minus_di).replace(0.0, np.nan)) * 100
    return dx.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()


# ==============================================================================
#  YES / NO / SAME
# ==============================================================================

def verdict(value: Optional[float], reference: Optional[float],
            tolerance_pct: float = 0.10) -> str:
    """
    The three answers that appear in the hand-built sheet.

    'Same' is used when the two numbers are close enough that calling it either
    way would be dishonest -- the same judgement call made by eye off a chart.
    Tolerance is a percentage of the reference value.
    """
    if value is None or reference is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(reference, float) and math.isnan(reference):
        return ""
    if reference == 0:
        return "Yes" if value > 0 else "No"
    diff_pct = (value - reference) / abs(reference) * 100.0
    if abs(diff_pct) <= tolerance_pct:
        return "Same"
    return "Yes" if diff_pct > 0 else "No"


def is_bull(answer: str) -> bool:
    return answer == "Yes"


def safe(x: Any) -> Optional[float]:
    """Turn NaN / numpy types into something JSON and Excel both accept."""
    if x is None:
        return None
    try:
        f = float(x)
    except (TypeError, ValueError):
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return f


# ==============================================================================
#  DATA DOWNLOAD
# ==============================================================================

def download_history(tickers: Sequence[str], period: str = "2y") -> Dict[str, pd.DataFrame]:
    """One batched request for every ticker, then split into per-ticker frames."""
    print(f"  Downloading {len(tickers)} tickers from Yahoo Finance ...")
    raw = yf.download(
        list(tickers), period=period, interval="1d",
        auto_adjust=False, progress=False, group_by="ticker", threads=True,
    )
    frames: Dict[str, pd.DataFrame] = {}
    for t in tickers:
        try:
            df = raw[t] if isinstance(raw.columns, pd.MultiIndex) else raw
            df = df.dropna(subset=["Close"]).copy()
            if len(df) >= 30:
                frames[t] = df
            else:
                print(f"    ! {t}: only {len(df)} rows of history, skipped")
        except Exception as exc:
            print(f"    ! {t}: {exc}")
    print(f"  Got usable history for {len(frames)} of {len(tickers)}.")
    return frames


def download_profiles(tickers: Sequence[str]) -> Dict[str, Dict[str, Any]]:
    """Company names, exchange and market cap -- cheap metadata, fetched in parallel."""
    def one(t: str) -> tuple:
        info: Dict[str, Any] = {"name": t, "exchange": "", "market_cap": None}
        try:
            tk = yf.Ticker(t)
            meta = tk.history_metadata or {}
            info["name"] = meta.get("longName") or meta.get("shortName") or t
            info["exchange"] = meta.get("fullExchangeName") or ""
            info["market_cap"] = safe(getattr(tk.fast_info, "market_cap", None))
        except Exception:
            pass
        return t, info

    print("  Fetching company names ...")
    out: Dict[str, Dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=8) as pool:
        for t, info in pool.map(one, tickers):
            out[t] = info
    return out


_TAG_RE = re.compile(r"<[^>]+>")


def _clean(text: str) -> str:
    text = _TAG_RE.sub("", text or "")
    text = (text.replace("&amp;", "&").replace("&quot;", '"')
                .replace("&#39;", "'").replace("&lt;", "<").replace("&gt;", ">"))
    return " ".join(text.split())


def download_news(tickers: Sequence[str], per_ticker: int = 3) -> Dict[str, List[Dict[str, str]]]:
    """
    Three headlines per ticker from Yahoo Finance's public RSS feed.
    Falls back to yfinance's own news list if the feed is unavailable.
    """
    import requests

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 (compatible; StockAnalysis/1.0)"})

    def one(t: str) -> tuple:
        items: List[Dict[str, str]] = []
        url = (f"https://feeds.finance.yahoo.com/rss/2.0/headline"
               f"?s={t}&region=US&lang=en-US")
        try:
            resp = session.get(url, timeout=12)
            if resp.ok:
                root = ET.fromstring(resp.content)
                for node in root.iter("item"):
                    title = _clean((node.findtext("title") or ""))
                    link = (node.findtext("link") or "").strip()
                    pub = (node.findtext("pubDate") or "").strip()
                    if title and link:
                        items.append({"title": title, "url": link, "published": pub,
                                      "source": _source_from(link)})
                    if len(items) >= per_ticker:
                        break
        except Exception:
            pass

        if not items:                                   # fallback
            try:
                for entry in (yf.Ticker(t).news or [])[:per_ticker]:
                    content = entry.get("content", entry)
                    title = content.get("title") or ""
                    link = ((content.get("canonicalUrl") or {}).get("url")
                            or (content.get("clickThroughUrl") or {}).get("url")
                            or entry.get("link") or "")
                    if title and link:
                        items.append({
                            "title": _clean(title), "url": link,
                            "published": content.get("pubDate", ""),
                            "source": _source_from(link),
                        })
            except Exception:
                pass
        return t, items[:per_ticker]

    print("  Fetching news headlines ...")
    out: Dict[str, List[Dict[str, str]]] = {}
    with ThreadPoolExecutor(max_workers=10) as pool:
        for t, items in pool.map(one, tickers):
            out[t] = items
    found = sum(1 for v in out.values() if v)
    print(f"  Headlines found for {found} of {len(tickers)} tickers.")
    return out


def _source_from(url: str) -> str:
    m = re.match(r"https?://(?:www\.)?([^/]+)", url or "")
    return m.group(1) if m else ""


# ==============================================================================
#  PER-TICKER ANALYSIS
# ==============================================================================

def analyse(ticker: str, df: pd.DataFrame, asof: pd.Timestamp,
            cfg: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Everything the workbook needs for one ticker on one day."""
    hist = df[df.index <= asof]
    if len(hist) < 30:
        return None

    close, high, low = hist["Close"], hist["High"], hist["Low"]
    open_, volume = hist["Open"], hist["Volume"]
    tol = cfg["same_tolerance_pct"]

    ema_f = ema(close, cfg["ema_fast"])
    ema_s = ema(close, cfg["ema_slow"])
    # The sheet has fixed columns for the 50, 100 and 200 day averages, so those
    # are always computed whatever else the watchlist asks for.
    sma_periods = sorted(set(cfg["sma_periods"]) | {50, 100, 200})
    smas = {p: sma(close, p) for p in sma_periods}

    fast, slow, sig_span = cfg["macd"]
    macd_line = ema(close, fast) - ema(close, slow)
    macd_signal = ema(macd_line, sig_span)
    macd_hist = macd_line - macd_signal

    bb_mid = sma(close, cfg["bb_period"])
    bb_sd = close.rolling(cfg["bb_period"], min_periods=cfg["bb_period"]).std(ddof=0)
    bb_up = bb_mid + cfg["bb_stdev"] * bb_sd
    bb_low = bb_mid - cfg["bb_stdev"] * bb_sd

    rsi_s = rsi(close, cfg["rsi_period"])
    atr_s = atr(high, low, close, cfg["atr_period"])
    stoch_s = stochastic_k(high, low, close, cfg["stoch_period"])
    adx_s = adx(high, low, close, 14)
    vol_ma = sma(volume.astype(float), cfg["volume_ma"])

    i = -1
    c = float(close.iloc[i])
    c_prev = float(close.iloc[i - 1])
    v = float(volume.iloc[i])
    v_prev = float(volume.iloc[i - 1])

    def back(series: pd.Series, n: int) -> Optional[float]:
        return safe(series.iloc[-1 - n]) if len(series) > n else None

    def ret(n: int) -> Optional[float]:
        if len(close) <= n:
            return None
        return c / float(close.iloc[-1 - n]) - 1.0

    # year-to-date, measured from the last close of the previous year
    ytd = None
    year_start = hist[hist.index.year < asof.year]
    if len(year_start):
        ytd = c / float(year_start["Close"].iloc[-1]) - 1.0

    window_52w = hist.tail(252)
    hi52 = float(window_52w["High"].max())
    lo52 = float(window_52w["Low"].min())

    a = {
        # --- his eleven columns -------------------------------------------
        "pct_change":  c / c_prev - 1.0,
        "c_gt_prev":   verdict(c, c_prev, tol),
        "v_gt_prev":   verdict(v, v_prev, 3.0),      # volume gets a wider band
        "macd_gt_sig": verdict(safe(macd_line.iloc[i]), safe(macd_signal.iloc[i]), 0.5),
        "c_gt_ema8":   verdict(c, safe(ema_f.iloc[i]), tol),
        "c_gt_ema21":  verdict(c, safe(ema_s.iloc[i]), tol),
        "c_gt_sma50":  verdict(c, back(smas[50], 0), tol),
        "c_gt_bbu":    verdict(c, safe(bb_up.iloc[i]), tol),
        "c_gt_bbl":    verdict(c, safe(bb_low.iloc[i]), tol),
        "c_gt_bbm":    verdict(c, safe(bb_mid.iloc[i]), tol),
        # --- price & volume -----------------------------------------------
        "last": c, "prev_close": c_prev, "chg": c - c_prev,
        "open": safe(open_.iloc[i]), "high": safe(high.iloc[i]), "low": safe(low.iloc[i]),
        "volume": v,
        "vol_ma20": safe(vol_ma.iloc[i]),
        "rvol": safe(v / vol_ma.iloc[i]) if safe(vol_ma.iloc[i]) else None,
        # --- momentum ------------------------------------------------------
        "rsi": safe(rsi_s.iloc[i]),
        "rsi_prev": back(rsi_s, 1),
        "macd": safe(macd_line.iloc[i]),
        "macd_signal": safe(macd_signal.iloc[i]),
        "macd_hist": safe(macd_hist.iloc[i]),
        "macd_hist_prev": back(macd_hist, 1),
        "stoch": safe(stoch_s.iloc[i]),
        "adx": safe(adx_s.iloc[i]),
        # --- moving averages -----------------------------------------------
        "ema8": safe(ema_f.iloc[i]), "ema21": safe(ema_s.iloc[i]),
        "ema8_prev": back(ema_f, 1), "ema21_prev": back(ema_s, 1),
        "sma50": back(smas[50], 0), "sma100": back(smas[100], 0),
        "sma200": back(smas[200], 0),
        "sma50_prev": back(smas[50], 1), "sma200_prev": back(smas[200], 1),
        "close_prev": c_prev,
        # --- bollinger ------------------------------------------------------
        "bb_up": safe(bb_up.iloc[i]), "bb_mid": safe(bb_mid.iloc[i]),
        "bb_low": safe(bb_low.iloc[i]),
        # --- range & returns -------------------------------------------------
        "atr": safe(atr_s.iloc[i]),
        "hi52": hi52, "lo52": lo52,
        "r5": ret(5), "r21": ret(21), "r63": ret(63), "r126": ret(126), "ytd": ytd,
        "date": asof.strftime("%Y-%m-%d"),
    }

    up, low_ = a["bb_up"], a["bb_low"]
    a["pct_b"] = safe((c - low_) / (up - low_)) if up and low_ and up != low_ else None
    a["bb_width"] = safe((up - low_) / a["bb_mid"] * 100) if up and low_ and a["bb_mid"] else None
    a["atr_pct"] = safe(a["atr"] / c * 100) if a["atr"] else None
    a["off_hi"] = safe((c / hi52 - 1) * 100) if hi52 else None
    a["range_pos"] = safe((c - lo52) / (hi52 - lo52) * 100) if hi52 != lo52 else None

    # ---- bull score: nine independent checks, each worth one point ---------
    checks = [
        c > c_prev,
        v > v_prev,
        (a["macd"] or 0) > (a["macd_signal"] or 0),
        a["ema8"] is not None and c > a["ema8"],
        a["ema21"] is not None and c > a["ema21"],
        a["sma50"] is not None and c > a["sma50"],
        a["sma200"] is not None and c > a["sma200"],
        a["bb_mid"] is not None and c > a["bb_mid"],
        (a["rsi"] or 0) > 50,
    ]
    a["score"] = int(sum(1 for x in checks if x))
    a["score_max"] = len(checks)
    a["verdict"] = score_label(a["score"])

    # ---- events worth a separate mention ----------------------------------
    events: List[str] = []
    if a["macd_hist"] is not None and a["macd_hist_prev"] is not None:
        if a["macd_hist"] > 0 >= a["macd_hist_prev"]:
            events.append("MACD crossed up")
        elif a["macd_hist"] < 0 <= a["macd_hist_prev"]:
            events.append("MACD crossed down")
    if a["ema8"] and a["ema21"] and a["ema8_prev"] and a["ema21_prev"]:
        if a["ema8"] > a["ema21"] and a["ema8_prev"] <= a["ema21_prev"]:
            events.append("8 EMA crossed above 21")
        elif a["ema8"] < a["ema21"] and a["ema8_prev"] >= a["ema21_prev"]:
            events.append("8 EMA crossed below 21")
    if a["bb_up"] and c > a["bb_up"]:
        events.append("Closed above upper band")
    if a["bb_low"] and c < a["bb_low"]:
        events.append("Closed below lower band")
    if a["rvol"] and a["rvol"] >= 1.75:
        events.append(f"Volume {a['rvol']:.1f}x its 20-day average")
    if a["rsi"] and a["rsi"] >= 70:
        events.append(f"RSI overbought ({a['rsi']:.0f})")
    if a["rsi"] and a["rsi"] <= 30:
        events.append(f"RSI oversold ({a['rsi']:.0f})")
    if a["rsi"] is not None and a["rsi_prev"] is not None:
        if a["rsi"] > 50 >= a["rsi_prev"]:
            events.append("RSI crossed back above 50")
        elif a["rsi"] < 50 <= a["rsi_prev"]:
            events.append("RSI dropped below 50")
    for period, label in ((50, "50 DMA"), (200, "200 DMA")):
        now_ma = a.get(f"sma{period}")
        was_ma = a.get(f"sma{period}_prev")
        if now_ma and was_ma:
            if c > now_ma and c_prev <= was_ma:
                events.append(f"Reclaimed the {label}")
            elif c < now_ma and c_prev >= was_ma:
                events.append(f"Lost the {label}")
    gap = (a["open"] / c_prev - 1) * 100 if a["open"] and c_prev else 0
    if gap >= 1.5:
        events.append(f"Gapped up {gap:.1f}% at the open")
    elif gap <= -1.5:
        events.append(f"Gapped down {abs(gap):.1f}% at the open")
    if a["bb_width"] is not None:
        width_hist = ((bb_up - bb_low) / bb_mid * 100).dropna().tail(120)
        if len(width_hist) >= 60 and a["bb_width"] <= width_hist.quantile(0.10):
            events.append("Bands squeezing — a big move often follows")
    if hi52 and c >= hi52 * 0.999:
        events.append("New 52-week high")
    if lo52 and c <= lo52 * 1.001:
        events.append("New 52-week low")
    a["events"] = events

    a["trend"] = trend_label(a)

    # A short price tail, for the sparklines on the website. Rounded hard so the
    # JSON the browser downloads stays small.
    tail = close.tail(90)
    a["spark"] = [round(float(x), 4) for x in tail]
    a["spark_dates"] = [d.strftime("%Y-%m-%d") for d in tail.index]
    a["spark_volume"] = [int(x) for x in volume.tail(90)]
    return a


def score_label(score: int) -> str:
    if score >= 8:
        return "Very Strong"
    if score >= 6:
        return "Strong"
    if score >= 4:
        return "Neutral"
    if score >= 2:
        return "Weak"
    return "Very Weak"


def trend_label(a: Dict[str, Any]) -> str:
    c = a["last"]
    mas = [a.get("ema8"), a.get("ema21"), a.get("sma50"), a.get("sma200")]
    mas = [m for m in mas if m]
    if not mas:
        return ""
    above = sum(1 for m in mas if c > m)
    if above == len(mas):
        return "Above every average"
    if above == 0:
        return "Below every average"
    if a.get("sma200") and c > a["sma200"]:
        return "Long-term up, short-term mixed"
    return "Long-term down, short-term mixed"


# ==============================================================================
#  COLUMN LAYOUT
# ==============================================================================
# (key, header, width, number-format, alignment)
# Columns A-M are byte-for-byte the original hand-built sheet.

COLUMNS: List[tuple] = [
    ("rank_num",   "#",                             4.5,  None,        "center"),
    ("ticker",     "Ticker",                        9.5,  None,        "left"),
    ("pct_change", "%change (From Watchlist)",     12.9,  "0.00%",     "center"),
    ("strength",   "Prev Strength (Based on Perf)", 12.6, "0",         "center"),
    ("c_gt_prev",  "Price\n>Pre-Day",               9.4,  None,        "center"),
    ("v_gt_prev",  "Yest Volum\n>2Pre-Day",        10.4,  None,        "center"),
    ("macd_gt_sig", "MACD\n(Green>Red)",           12.7,  None,        "center"),
    ("c_gt_ema8",  "Price >8EMA",                   9.5,  None,        "center"),
    ("c_gt_ema21", "> 21EMA",                       9.0,  None,        "center"),
    ("c_gt_sma50", "Price > 50DMA",                11.5,  None,        "center"),
    ("c_gt_bbu",   "Price > Upper BB",             11.5,  None,        "center"),
    ("c_gt_bbl",   "Price > Lower BB",             11.5,  None,        "center"),
    ("c_gt_bbm",   "Price > Mid BB",               11.5,  None,        "center"),
    # ---------------- everything below is new -----------------------------
    ("score",      "Bull\nScore",                   7.5,  "0",         "center"),
    ("verdict",    "Signal",                       13.0,  None,        "center"),
    ("trend",      "Trend",                        26.0,  None,        "left"),
    ("name",       "Company / Fund",               32.0,  None,        "left"),
    ("last",       "Last",                         10.0,  "#,##0.00",  "right"),
    ("chg",        "Chg $",                         9.5,  "+#,##0.00;-#,##0.00;0.00", "right"),
    ("open",       "Open",                          9.5,  "#,##0.00",  "right"),
    ("high",       "High",                          9.5,  "#,##0.00",  "right"),
    ("low",        "Low",                           9.5,  "#,##0.00",  "right"),
    ("volume",     "Volume",                       13.0,  "#,##0",     "right"),
    ("rvol",       "Vol vs\n20d Avg",               9.5,  '0.00"x"',   "center"),
    ("rsi",        "RSI (14)",                      9.0,  "0.0",       "center"),
    ("stoch",      "Stoch %K",                      9.5,  "0.0",       "center"),
    ("adx",        "ADX (14)",                      9.0,  "0.0",       "center"),
    ("macd",       "MACD",                         10.0,  "0.000",     "right"),
    ("macd_signal", "Signal",                      10.0,  "0.000",     "right"),
    ("macd_hist",  "Histogram",                    10.5,  "0.000",     "right"),
    ("ema8",       "8 EMA",                        10.5,  "#,##0.00",  "right"),
    ("ema21",      "21 EMA",                       10.5,  "#,##0.00",  "right"),
    ("sma50",      "50 DMA",                       10.5,  "#,##0.00",  "right"),
    ("sma100",     "100 DMA",                      10.5,  "#,##0.00",  "right"),
    ("sma200",     "200 DMA",                      10.5,  "#,##0.00",  "right"),
    ("bb_up",      "BB Upper",                     10.5,  "#,##0.00",  "right"),
    ("bb_mid",     "BB Mid",                       10.5,  "#,##0.00",  "right"),
    ("bb_low",     "BB Lower",                     10.5,  "#,##0.00",  "right"),
    ("pct_b",      "%B",                            8.0,  "0.00",      "center"),
    ("bb_width",   "Band\nWidth",                   9.0,  '0.0"%"',    "center"),
    ("atr_pct",    "ATR (14)",                      9.5,  '0.00"%"',   "center"),
    ("hi52",       "52W High",                     10.5,  "#,##0.00",  "right"),
    ("lo52",       "52W Low",                      10.5,  "#,##0.00",  "right"),
    ("off_hi",     "% Off\n52W High",              10.0,  '0.0"%"',    "center"),
    ("range_pos",  "52W Range\nPosition",          11.0,  '0"%"',      "center"),
    ("r5",         "5-Day",                         9.0,  "0.00%",     "center"),
    ("r21",        "1-Month",                       9.5,  "0.00%",     "center"),
    ("r63",        "3-Month",                       9.5,  "0.00%",     "center"),
    ("ytd",        "Year to Date",                 10.5,  "0.00%",     "center"),
    ("news1",      "Top News 1",                   52.0,  None,        "left"),
    ("news2",      "Top News 2",                   52.0,  None,        "left"),
    ("news3",      "Top News 3",                   52.0,  None,        "left"),
]

COL_INDEX = {key: n + 1 for n, (key, *_rest) in enumerate(COLUMNS)}

# Group bands drawn across row 5.
BANDS: List[tuple] = [
    ("pct_change",  "c_gt_bbm",   "THE DAILY CHECKLIST", "1E3A5F"),
    ("score",       "trend",      "VERDICT",             "0F5132"),
    ("name",        "rvol",       "PRICE & VOLUME",      "1F2937"),
    ("rsi",         "macd_hist",  "MOMENTUM",            "5B21B6"),
    ("ema8",        "sma200",     "MOVING AVERAGES",     "0C4A6E"),
    ("bb_up",       "atr_pct",    "BOLLINGER & VOLATILITY", "7C2D12"),
    ("hi52",        "ytd",        "RANGE & RETURNS",     "334155"),
    ("news1",       "news3",      "TOP NEWS  (click a headline to open it)", "9A3412"),
]

YESNO_KEYS = ["c_gt_prev", "v_gt_prev", "macd_gt_sig", "c_gt_ema8", "c_gt_ema21",
              "c_gt_sma50", "c_gt_bbu", "c_gt_bbl", "c_gt_bbm"]


# ==============================================================================
#  EXCEL HELPERS
# ==============================================================================

THIN = Side(style="thin", color=BORDER_GREY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def put(ws: Worksheet, row: int, col: int, value: Any, *, bold: bool = False,
        size: float = 10.5, colour: str = "1E293B", fill: Optional[str] = None,
        fmt: Optional[str] = None, align: str = "left", wrap: bool = False,
        border: bool = True, link: Optional[str] = None, italic: bool = False):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = Font(name="Aptos Narrow", size=size, bold=bold, italic=italic,
                     color=colour, underline="single" if link else None)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        cell.border = BOX
    if link:
        cell.hyperlink = link
    return cell


def title_block(ws: Worksheet, asof_label: str, generated: str, settings: Dict[str, Any]):
    """Rows 1-3: the chart link, the run stamp and the indicator settings."""
    put(ws, 1, 1, CHART_URL, link=CHART_URL, colour="0563C1", size=10.5, border=False)
    put(ws, 2, 1, f"Trading day: {asof_label}", bold=True, size=11, colour=INK, border=False)
    put(ws, 3, 1, f"Built {generated} · Yahoo Finance data",
        size=9, colour=LABEL_GREY, italic=True, border=False)

    legend = [
        ("EMA:", f"{settings['ema_fast']}, {settings['ema_slow']}"),
        ("DMA:", ", ".join(str(p) for p in settings["sma_periods"])),
        ("BB:",  f"{settings['bb_period']}, {settings['bb_stdev']}"),
    ]
    for n, (label, value) in enumerate(legend):
        put(ws, 1 + n, 8, label, bold=True, size=10.5, colour=INK,
            align="right", border=False)
        put(ws, 1 + n, 10, value, size=10.5, colour="334155", border=False)

    extra = [
        ("MACD:",  f"{settings['macd'][0]}, {settings['macd'][1]}, {settings['macd'][2]}"),
        ("Volume MA:", str(settings["volume_ma"])),
        ("RSI / ATR:", f"{settings['rsi_period']}, {settings['atr_period']}"),
    ]
    for n, (label, value) in enumerate(extra):
        put(ws, 1 + n, 16, label, bold=True, size=10.5, colour=INK,
            align="right", border=False)
        put(ws, 1 + n, 17, value, size=10.5, colour="334155", border=False)


def market_box(ws: Worksheet, rows: List[Dict[str, Any]]):
    """A small at-a-glance panel in the top right of the day sheet."""
    live = [r for r in rows if r.get("a")]
    if not live:
        return
    up = sum(1 for r in live if (r["a"]["pct_change"] or 0) > 0)
    down = sum(1 for r in live if (r["a"]["pct_change"] or 0) < 0)
    avg = sum(r["a"]["score"] for r in live) / len(live)
    strong = sum(1 for r in live if r["a"]["score"] >= 6)

    stats = [
        ("Advancing",   f"{up} of {len(live)}", GREEN_FG),
        ("Declining",   f"{down} of {len(live)}", RED_FG),
        ("Average score", f"{avg:.1f} / 9", INK),
        ("Scoring 6+",  f"{strong} of {len(live)}", BLUE_FG),
    ]
    start = COL_INDEX["rsi"]
    put(ws, 1, start, "TODAY AT A GLANCE", bold=True, size=10,
        colour="FFFFFF", fill=INK, align="center", border=False)
    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + 7)
    for n, (label, value, colour) in enumerate(stats):
        c = start + n * 2
        put(ws, 2, c, label, size=9, colour=LABEL_GREY, align="center", border=False)
        put(ws, 3, c, value, bold=True, size=11, colour=colour, align="center", border=False)
        ws.merge_cells(start_row=2, start_column=c, end_row=2, end_column=c + 1)
        ws.merge_cells(start_row=3, start_column=c, end_row=3, end_column=c + 1)


def header_rows(ws: Worksheet):
    """Row 5 group bands, row 6 the column headers themselves."""
    for first, last, label, colour in BANDS:
        c1, c2 = COL_INDEX[first], COL_INDEX[last]
        ws.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
        put(ws, 5, c1, label, bold=True, size=9.5, colour="FFFFFF",
            fill=colour, align="center", border=False)
        for c in range(c1, c2 + 1):
            ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor=colour)

    for n, (key, header, _w, _f, align) in enumerate(COLUMNS, start=1):
        put(ws, 6, n, header, bold=True, size=10, colour="FFFFFF", fill=INK,
            align="center" if align != "left" else "left", wrap=True)
    ws.row_dimensions[5].height = 17
    ws.row_dimensions[6].height = 34
    for n, (_k, _h, width, *_r) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(n)].width = width


def conditional_formats(ws: Worksheet, first: int, last: int,
                        blocks: Optional[List[tuple]] = None):
    """Live rules, so the colours follow along if a cell is edited by hand."""
    if last < first:
        return

    def rng(key: str) -> str:
        col = get_column_letter(COL_INDEX[key])
        return f"{col}{first}:{col}{last}"

    yes_no_span = (f"{get_column_letter(COL_INDEX['c_gt_prev'])}{first}:"
                   f"{get_column_letter(COL_INDEX['c_gt_bbm'])}{last}")
    ws.conditional_formatting.add(yes_no_span, CellIsRule(
        operator="equal", formula=['"Yes"'],
        fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"),
        font=Font(name="Aptos Narrow", size=10.5, bold=True, color=GREEN_FG)))
    ws.conditional_formatting.add(yes_no_span, CellIsRule(
        operator="equal", formula=['"No"'],
        fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"),
        font=Font(name="Aptos Narrow", size=10.5, bold=True, color=RED_FG)))
    ws.conditional_formatting.add(yes_no_span, CellIsRule(
        operator="equal", formula=['"Same"'],
        fill=PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid"),
        font=Font(name="Aptos Narrow", size=10.5, bold=True, color=AMBER_FG)))

    red_green = dict(start_type="num", start_value=-0.03, start_color="F8B4B4",
                     mid_type="num", mid_value=0, mid_color="FFFFFF",
                     end_type="num", end_value=0.03, end_color="8FD9A8")
    for key in ("pct_change", "r5", "r21", "r63", "ytd"):
        ws.conditional_formatting.add(rng(key), ColorScaleRule(**red_green))

    ws.conditional_formatting.add(rng("chg"), ColorScaleRule(
        start_type="percentile", start_value=5, start_color="F8B4B4",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="percentile", end_value=95, end_color="8FD9A8"))

    ws.conditional_formatting.add(rng("score"), DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=9,
        color="2FA45C", showValue=True, minLength=None, maxLength=None))

    # Rank is graded inside each block, because being 11th of eleven sectors and
    # 11th of twenty-five stocks are not the same thing.
    strength_col = get_column_letter(COL_INDEX["strength"])
    for block_first, block_last in (blocks or [(first, last)]):
        ws.conditional_formatting.add(
            f"{strength_col}{block_first}:{strength_col}{block_last}",
            ColorScaleRule(start_type="min", start_color="8FD9A8",
                           mid_type="percentile", mid_value=50, mid_color="FFF7C2",
                           end_type="max", end_color="F8B4B4"))

    ws.conditional_formatting.add(rng("rsi"), ColorScaleRule(
        start_type="num", start_value=30, start_color="8FD9A8",
        mid_type="num", mid_value=50, mid_color="FFFFFF",
        end_type="num", end_value=70, end_color="F8B4B4"))

    ws.conditional_formatting.add(rng("stoch"), ColorScaleRule(
        start_type="num", start_value=20, start_color="8FD9A8",
        mid_type="num", mid_value=50, mid_color="FFFFFF",
        end_type="num", end_value=80, end_color="F8B4B4"))

    ws.conditional_formatting.add(rng("rvol"), ColorScaleRule(
        start_type="num", start_value=0.5, start_color="FFFFFF",
        mid_type="num", mid_value=1.0, mid_color="DDEBFB",
        end_type="num", end_value=3.0, end_color="5FA8E8"))

    ws.conditional_formatting.add(rng("adx"), ColorScaleRule(
        start_type="num", start_value=10, start_color="FFFFFF",
        end_type="num", end_value=40, end_color="C4B5FD"))

    ws.conditional_formatting.add(rng("pct_b"), ColorScaleRule(
        start_type="num", start_value=0, start_color="8FD9A8",
        mid_type="num", mid_value=0.5, mid_color="FFFFFF",
        end_type="num", end_value=1, end_color="F8B4B4"))

    ws.conditional_formatting.add(rng("off_hi"), ColorScaleRule(
        start_type="num", start_value=-40, start_color="F8B4B4",
        mid_type="num", mid_value=-15, mid_color="FFF7C2",
        end_type="num", end_value=0, end_color="8FD9A8"))

    ws.conditional_formatting.add(rng("range_pos"), ColorScaleRule(
        start_type="num", start_value=0, start_color="F8B4B4",
        mid_type="num", mid_value=50, mid_color="FFF7C2",
        end_type="num", end_value=100, end_color="8FD9A8"))

    for key in ("macd_hist",):
        ws.conditional_formatting.add(rng(key), ColorScaleRule(
            start_type="percentile", start_value=5, start_color="F8B4B4",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="percentile", end_value=95, end_color="8FD9A8"))

    verdict_col = f"{get_column_letter(COL_INDEX['verdict'])}{first}:" \
                  f"{get_column_letter(COL_INDEX['verdict'])}{last}"
    for text, bg, fg in (("Very Strong", "BBF0CE", "0B5C2E"),
                         ("Strong", GREEN_BG, GREEN_FG),
                         ("Neutral", "EEF2F7", "475569"),
                         ("Weak", AMBER_BG, AMBER_FG),
                         ("Very Weak", RED_BG, RED_FG)):
        ws.conditional_formatting.add(verdict_col, CellIsRule(
            operator="equal", formula=[f'"{text}"'],
            fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
            font=Font(name="Aptos Narrow", size=10.5, bold=True, color=fg)))


# ==============================================================================
#  THE DAY SHEET
# ==============================================================================

def build_day_sheet(wb: Workbook, sheet_name: str, rows: List[Dict[str, Any]],
                    asof_label: str, generated: str, settings: Dict[str, Any],
                    news: Dict[str, List[Dict[str, str]]], profiles: Dict[str, Dict]):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = ACCENT
    title_block(ws, asof_label, generated, settings)
    market_box(ws, rows)
    header_rows(ws)

    # The original sheet put sectors on rows 7-17, the indexes on 19-22 and the
    # watchlist from 24 down, with one blank row between each block. Those exact
    # positions are kept here -- the section labels go into the blank rows that
    # were already there, and the first label sits on row 4, above the headers.
    r = 7
    first_data = r
    blocks: List[tuple] = []

    for n_group, (gkey, gtitle, items) in enumerate(rows_by_group(rows)):
        # The first block needs no spacer -- its label sits on row 4. Every later
        # block puts its label on the blank row that separated the blocks anyway.
        section_divider(ws, 4 if n_group == 0 else r, gkey, gtitle, len(items))
        if n_group > 0:
            r += 1
        block_first = r
        for n, item in enumerate(items, start=1):
            write_row(ws, r, n, item, news, profiles, zebra=(n % 2 == 0),
                      accent=SECTION_FILLS.get(gkey, SUBTLE))
            r += 1
        blocks.append((block_first, r - 1))

    last_data = r - 1
    conditional_formats(ws, first_data, last_data, blocks)

    ws.freeze_panes = "C7"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = (f"A6:{get_column_letter(len(COLUMNS))}{last_data}")
    ws.print_title_rows = "5:6"
    return ws


SECTION_INK = {"sectors": "1E3A5F", "indices": "0F5132", "stocks": "5B21B6"}


def section_divider(ws: Worksheet, row: int, gkey: str, title: str, count: int):
    """
    A slim coloured bar naming the block of tickers underneath it.

    Deliberately unmerged: this row sits inside the auto-filter range, and a
    merged range in there makes Excel offer to repair the file on open. The
    label is written into column B and simply overflows across the bar.
    """
    ink = SECTION_INK.get(gkey, INK)
    put(ws, row, 1, None, fill=ink, border=False)
    put(ws, row, 2, f"{title}   ({count})", bold=True, size=9.5, colour="FFFFFF",
        fill=ink, align="left", border=False)
    for c in range(3, COL_INDEX["verdict"] + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=ink)
    ws.row_dimensions[row].height = 14


def rows_by_group(rows: List[Dict[str, Any]]):
    out: List[tuple] = []
    for row in rows:
        if not out or out[-1][0] != row["group"]:
            out.append((row["group"], row["group_title"], []))
        out[-1][2].append(row)
    return out


def write_row(ws: Worksheet, r: int, n: int, item: Dict[str, Any],
              news: Dict[str, List[Dict[str, str]]], profiles: Dict[str, Dict],
              zebra: bool, accent: str = SUBTLE):
    ticker = item["ticker"]
    a = item.get("a") or {}
    bg = SUBTLE if zebra else None

    values: Dict[str, Any] = {
        "rank_num": n if item.get("numbered") else None,
        "ticker": ticker,
        "strength": item.get("strength"),
        "name": profiles.get(ticker, {}).get("name", ticker),
    }
    for key in ("pct_change", "score", "verdict", "trend", "last", "chg", "open",
                "high", "low", "volume", "rvol", "rsi", "stoch", "adx", "macd",
                "macd_signal", "macd_hist", "ema8", "ema21", "sma50", "sma100",
                "sma200", "bb_up", "bb_mid", "bb_low", "pct_b", "bb_width",
                "atr_pct", "hi52", "lo52", "off_hi", "range_pos",
                "r5", "r21", "r63", "ytd", *YESNO_KEYS):
        values[key] = a.get(key)

    for key, header, _w, fmt, align in COLUMNS:
        if key.startswith("news"):
            continue
        col = COL_INDEX[key]
        value = values.get(key)
        bold = key in ("ticker", "score", "last")
        size = 11 if key == "ticker" else 10.5
        colour = INK if key == "ticker" else "1E293B"
        fill = accent if key in ("rank_num", "ticker") else bg
        put(ws, r, col, value, fmt=fmt, align=align, bold=bold, size=size,
            colour=colour, fill=fill)

    headlines = news.get(ticker, [])
    for k in range(3):
        col = COL_INDEX[f"news{k + 1}"]
        if k < len(headlines):
            h = headlines[k]
            label = h["title"]
            if len(label) > 110:
                label = label[:107] + "..."
            put(ws, r, col, label, link=h["url"], colour="0563C1",
                size=9.5, align="left", fill=bg)
            ws.cell(row=r, column=col).alignment = Alignment(
                horizontal="left", vertical="center", shrink_to_fit=False)
        else:
            put(ws, r, col, None, fill=bg)

    ws.row_dimensions[r].height = 17


# ==============================================================================
#  DASHBOARD SHEET
# ==============================================================================

def build_dashboard(wb: Workbook, rows: List[Dict[str, Any]], asof_label: str,
                    generated: str, settings: Dict[str, Any]):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0B1220"
    for col, width in zip("ABCDEFGHIJKLMNOPQR",
                          [3, 15, 34, 11, 11, 11, 11, 11, 11, 11, 3,
                           15, 30, 11, 11, 11, 11, 11]):
        ws.column_dimensions[col].width = width

    live = [r for r in rows if r.get("a")]

    # ---- banner -----------------------------------------------------------
    ws.merge_cells("B2:R4")
    put(ws, 2, 2, f"DAILY STOCK ANALYSIS   ·   {asof_label}", bold=True, size=22,
        colour="FFFFFF", fill=INK, align="left", border=False)
    for row in range(2, 5):
        for col in range(2, 19):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=INK)
    ws.merge_cells("B5:R5")
    put(ws, 5, 2, f"Built automatically {generated} · Yahoo Finance data · "
                  f"EMA {settings['ema_fast']}/{settings['ema_slow']} · "
                  f"DMA {'/'.join(str(p) for p in settings['sma_periods'])} · "
                  f"BB {settings['bb_period']},{settings['bb_stdev']} · "
                  f"MACD {'/'.join(str(p) for p in settings['macd'])}",
        size=9.5, colour="FFFFFF", fill="1E3A5F", align="left", border=False)
    for col in range(2, 19):
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor="1E3A5F")
    ws.row_dimensions[2].height = 12
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 12

    # ---- headline tiles ---------------------------------------------------
    up = sum(1 for r in live if (r["a"]["pct_change"] or 0) > 0)
    down = sum(1 for r in live if (r["a"]["pct_change"] or 0) < 0)
    avg_score = sum(r["a"]["score"] for r in live) / max(len(live), 1)
    breadth = up / max(up + down, 1) * 100
    above200 = sum(1 for r in live
                   if r["a"].get("sma200") and r["a"]["last"] > r["a"]["sma200"])
    events_n = sum(len(r["a"].get("events", [])) for r in live)

    tiles = [
        ("ADVANCING",       f"{up}",  f"of {len(live)} names",  GREEN_FG, GREEN_BG),
        ("DECLINING",       f"{down}", f"of {len(live)} names", RED_FG,   RED_BG),
        ("BREADTH",         f"{breadth:.0f}%", "advancing share",
         GREEN_FG if breadth >= 50 else RED_FG, GREEN_BG if breadth >= 50 else RED_BG),
        ("AVERAGE SCORE",   f"{avg_score:.1f}", "out of 9",     BLUE_FG,  BLUE_BG),
        ("ABOVE 200 DMA",   f"{above200}", f"of {len(live)} names", BLUE_FG, BLUE_BG),
        ("SIGNALS TODAY",   f"{events_n}", "see Signals tab",   AMBER_FG, AMBER_BG),
    ]
    row = 7
    for n, (label, big, sub, fg, bg) in enumerate(tiles):
        c = 2 + n * 3
        ws.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c + 2)
        ws.merge_cells(start_row=row + 1, start_column=c, end_row=row + 1, end_column=c + 2)
        ws.merge_cells(start_row=row + 2, start_column=c, end_row=row + 2, end_column=c + 2)
        put(ws, row, c, label, bold=True, size=8.5, colour=LABEL_GREY,
            fill=bg, align="center", border=False)
        put(ws, row + 1, c, big, bold=True, size=24, colour=fg, fill=bg,
            align="center", border=False)
        put(ws, row + 2, c, sub, size=9, colour=LABEL_GREY, fill=bg,
            align="center", border=False)
        for rr in (row, row + 1, row + 2):
            for cc in range(c, c + 3):
                ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=bg)
    ws.row_dimensions[row].height = 15
    ws.row_dimensions[row + 1].height = 32
    ws.row_dimensions[row + 2].height = 15

    # ---- sector strength table -------------------------------------------
    start = 12
    section_title(ws, start, 2, "SECTOR STRENGTH — today, ranked", span=9)
    sectors = [r for r in live if r["group"] == "sectors"]
    sectors.sort(key=lambda r: r["a"]["pct_change"] or 0, reverse=True)
    heads = ["Rank", "Sector ETF", "Name", "% Change", "Score", "RSI",
             "vs 50 DMA", "5-Day", "1-Month"]
    for n, h in enumerate(heads):
        put(ws, start + 1, 2 + n, h, bold=True, size=9.5, colour="FFFFFF",
            fill=INK, align="center")
    for n, r in enumerate(sectors):
        a = r["a"]
        rr = start + 2 + n
        vals = [n + 1, r["ticker"], r.get("name", ""), a["pct_change"], a["score"],
                a["rsi"],
                safe((a["last"] / a["sma50"] - 1) * 100) if a.get("sma50") else None,
                a["r5"], a["r21"]]
        fmts = [None, None, None, "0.00%", "0", "0.0", '0.0"%"', "0.00%", "0.00%"]
        aligns = ["center", "center", "left", "center", "center", "center",
                  "center", "center", "center"]
        for k, (v, f, al) in enumerate(zip(vals, fmts, aligns)):
            put(ws, rr, 2 + k, v, fmt=f, align=al,
                bold=(k == 1), fill=SUBTLE if n % 2 else None)
    end = start + 1 + len(sectors)
    ws.conditional_formatting.add(f"E{start + 2}:E{end}", ColorScaleRule(
        start_type="num", start_value=-0.02, start_color="F8B4B4",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=0.02, end_color="8FD9A8"))
    ws.conditional_formatting.add(f"F{start + 2}:F{end}", DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=9,
        color="2FA45C", showValue=True))
    for letter in ("I", "J"):
        ws.conditional_formatting.add(f"{letter}{start + 2}:{letter}{end}", ColorScaleRule(
            start_type="percentile", start_value=10, start_color="F8B4B4",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="percentile", end_value=90, end_color="8FD9A8"))

    chart = BarChart()
    chart.type = "bar"
    chart.style = 2
    chart.title = "Sector performance today"
    chart.y_axis.numFmt = "0.0%"
    chart.height, chart.width = 9.5, 12
    chart.add_data(Reference(ws, min_col=5, min_row=start + 1, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=3, min_row=start + 2, max_row=end))
    chart.legend = None
    ws.add_chart(chart, f"M{start + 1}")

    # ---- leaders and laggards --------------------------------------------
    start2 = end + 3
    ranked = sorted(live, key=lambda r: (r["a"]["score"], r["a"]["pct_change"] or 0),
                    reverse=True)
    two_tables(ws, start2, "STRONGEST TODAY", ranked[:10],
               "WEAKEST TODAY", list(reversed(ranked))[:10])

    movers = sorted(live, key=lambda r: r["a"]["pct_change"] or 0, reverse=True)
    start3 = start2 + 14
    two_tables(ws, start3, "BIGGEST GAINERS", movers[:10],
               "BIGGEST LOSERS", list(reversed(movers))[:10])

    ws.sheet_view.zoomScale = 90
    return ws


def section_title(ws: Worksheet, row: int, col: int, text: str, span: int = 8):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    put(ws, row, col, text, bold=True, size=12, colour=INK, align="left", border=False)
    ws.row_dimensions[row].height = 22


def two_tables(ws: Worksheet, row: int, left_title: str, left: List[Dict],
               right_title: str, right: List[Dict]):
    for offset, (title, items) in enumerate(((left_title, left), (right_title, right))):
        c = 2 + offset * 6
        section_title(ws, row, c, title, span=5)
        for n, h in enumerate(["Ticker", "% Change", "Score", "Signal", "RSI"]):
            put(ws, row + 1, c + n, h, bold=True, size=9.5, colour="FFFFFF",
                fill=INK, align="center")
        for n, r in enumerate(items):
            a = r["a"]
            rr = row + 2 + n
            cells = [(r["ticker"], None, "center", True),
                     (a["pct_change"], "0.00%", "center", False),
                     (a["score"], "0", "center", False),
                     (a["verdict"], None, "center", False),
                     (a["rsi"], "0.0", "center", False)]
            for k, (v, f, al, b) in enumerate(cells):
                put(ws, rr, c + k, v, fmt=f, align=al, bold=b,
                    fill=SUBTLE if n % 2 else None)
        last = row + 1 + len(items)
        col_pct = get_column_letter(c + 1)
        ws.conditional_formatting.add(f"{col_pct}{row + 2}:{col_pct}{last}", ColorScaleRule(
            start_type="num", start_value=-0.03, start_color="F8B4B4",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=0.03, end_color="8FD9A8"))
        col_score = get_column_letter(c + 2)
        ws.conditional_formatting.add(f"{col_score}{row + 2}:{col_score}{last}", DataBarRule(
            start_type="num", start_value=0, end_type="num", end_value=9,
            color="2FA45C", showValue=True))


# ==============================================================================
#  SECTOR ROTATION SHEET
# ==============================================================================

def build_rotation(wb: Workbook, rows: List[Dict[str, Any]], asof_label: str):
    ws = wb.create_sheet("Sector Rotation")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "5B21B6"
    widths = [3, 11, 34, 11, 11, 11, 11, 11, 11, 11, 11, 13, 13]
    for col, w in zip("ABCDEFGHIJKLM", widths):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:M2")
    put(ws, 2, 2, f"SECTOR ROTATION  ·  {asof_label}", bold=True, size=18,
        colour="FFFFFF", fill=INK, align="left", border=False)
    for c in range(2, 14):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=INK)
    ws.row_dimensions[2].height = 30
    put(ws, 3, 2, "Money moves between sectors. Green columns are where it is going.",
        size=10, colour=LABEL_GREY, italic=True, border=False)

    periods = [("pct_change", "Today"), ("r5", "1 Week"), ("r21", "1 Month"),
               ("r63", "3 Months"), ("r126", "6 Months"), ("ytd", "Year to Date")]
    heads = ["Ticker", "Sector"] + [label for _k, label in periods] + \
            ["Score", "Trend"]
    for n, h in enumerate(heads):
        put(ws, 5, 2 + n, h, bold=True, size=10, colour="FFFFFF", fill=INK,
            align="center", wrap=True)
    ws.row_dimensions[5].height = 26

    sectors = [r for r in rows if r["group"] == "sectors" and r.get("a")]
    sectors.sort(key=lambda r: r["a"].get("r21") or -9, reverse=True)
    for n, r in enumerate(sectors):
        a = r["a"]
        rr = 6 + n
        put(ws, rr, 2, r["ticker"], bold=True, align="center",
            fill=SUBTLE if n % 2 else None)
        put(ws, rr, 3, r.get("name", ""), align="left", fill=SUBTLE if n % 2 else None)
        for k, (key, _label) in enumerate(periods):
            put(ws, rr, 4 + k, a.get(key), fmt="0.00%", align="center",
                fill=SUBTLE if n % 2 else None)
        put(ws, rr, 4 + len(periods), a["score"], fmt="0", align="center",
            fill=SUBTLE if n % 2 else None)
        put(ws, rr, 5 + len(periods), a.get("trend", ""), align="left", size=9.5,
            fill=SUBTLE if n % 2 else None)

    last = 5 + len(sectors)
    for k in range(len(periods)):
        letter = get_column_letter(4 + k)
        ws.conditional_formatting.add(f"{letter}6:{letter}{last}", ColorScaleRule(
            start_type="min", start_color="F8B4B4",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="8FD9A8"))
    score_letter = get_column_letter(4 + len(periods))
    ws.conditional_formatting.add(f"{score_letter}6:{score_letter}{last}", DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=9,
        color="2FA45C", showValue=True))

    put(ws, last + 2, 2,
        "Read it left to right: a sector green in every column is in a real trend. "
        "Green today but red over a month is a bounce, not a trend.",
        size=10, colour=LABEL_GREY, italic=True, border=False)
    return ws


# ==============================================================================
#  SIGNALS SHEET
# ==============================================================================

def build_signals(wb: Workbook, rows: List[Dict[str, Any]], asof_label: str):
    ws = wb.create_sheet("Signals")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "B45309"
    for col, w in zip("ABCDEFG", [3, 11, 30, 40, 12, 11, 13]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:G2")
    put(ws, 2, 2, f"WHAT CHANGED TODAY  ·  {asof_label}", bold=True, size=18,
        colour="FFFFFF", fill=INK, align="left", border=False)
    for c in range(2, 8):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=INK)
    ws.row_dimensions[2].height = 30
    put(ws, 3, 2, "Only the tickers where something actually happened. "
                  "Everything else was quiet.",
        size=10, colour=LABEL_GREY, italic=True, border=False)

    for n, h in enumerate(["Ticker", "Name", "What happened", "% Change",
                           "Score", "Signal"]):
        put(ws, 5, 2 + n, h, bold=True, size=10, colour="FFFFFF", fill=INK,
            align="center")

    bullish = ("crossed up", "crossed above", "above upper", "52-week high", "oversold")
    r = 6
    entries = []
    for row in rows:
        a = row.get("a")
        if not a:
            continue
        for event in a.get("events", []):
            entries.append((row, event))
    entries.sort(key=lambda e: (e[0]["a"]["score"], e[0]["a"]["pct_change"] or 0),
                 reverse=True)

    for n, (row, event) in enumerate(entries):
        a = row["a"]
        good = any(k in event.lower() for k in bullish)
        bg = SUBTLE if n % 2 else None
        put(ws, r, 2, row["ticker"], bold=True, align="center", fill=bg)
        put(ws, r, 3, row.get("name", ""), align="left", size=9.5, fill=bg)
        put(ws, r, 4, event, align="left",
            colour=GREEN_FG if good else RED_FG, bold=True, fill=bg)
        put(ws, r, 5, a["pct_change"], fmt="0.00%", align="center", fill=bg)
        put(ws, r, 6, a["score"], fmt="0", align="center", fill=bg)
        put(ws, r, 7, a["verdict"], align="center", fill=bg)
        r += 1

    if not entries:
        put(ws, 6, 2, "Nothing crossed, broke out or spiked today. A quiet session.",
            italic=True, colour=LABEL_GREY, border=False)
    else:
        ws.conditional_formatting.add(f"E6:E{r - 1}", ColorScaleRule(
            start_type="num", start_value=-0.05, start_color="F8B4B4",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=0.05, end_color="8FD9A8"))
        ws.auto_filter.ref = f"B5:G{r - 1}"
    ws.freeze_panes = "B6"
    return ws


# ==============================================================================
#  GUIDE SHEET
# ==============================================================================

GUIDE: List[tuple] = [
    ("THE DAILY CHECKLIST", "", ""),
    ("%change", "The day's move, close against yesterday's close.",
     "Green is up, red is down. This is the same number the watchlist shows."),
    ("Prev Strength", "Rank of that move inside its own group. 1 is the strongest.",
     "Sectors are ranked against the other ten sectors, not against the stocks."),
    ("Price > Pre-Day", "Did it close higher than yesterday?",
     "Yes, No, or Same when the two closes are within a tenth of a percent."),
    ("Yest Volum > 2Pre-Day", "Did more shares trade than the day before?",
     "Rising volume behind a rising price is the healthy combination."),
    ("MACD (Green > Red)", "Is the MACD line above its signal line?",
     "Above means momentum is with the buyers. The Histogram column is the gap."),
    ("Price > 8 EMA", "Above the 8-day exponential average.",
     "The fastest line. Losing it first is usually the first crack."),
    ("> 21 EMA", "Above the 21-day exponential average.",
     "The line that separates a dip from a change of trend."),
    ("Price > 50 DMA", "Above the 50-day simple average.",
     "The line most desks watch for the medium-term trend."),
    ("Price > Upper BB", "Above the top Bollinger Band.",
     "Rare. Either a genuine breakout or a stretched move about to snap back."),
    ("Price > Lower BB", "Above the bottom Bollinger Band.",
     "Almost always Yes. A No here means something is badly wrong."),
    ("Price > Mid BB", "Above the middle band, which is the 20-day average.",
     "The simplest read on whether the last month has been up or down."),
    ("", "", ""),
    ("THE NEW COLUMNS", "", ""),
    ("Bull Score", "Nine checks, one point each, added up.",
     "8 or 9 is a strong tape. 0 or 1 is a broken one. The bar length shows it."),
    ("Signal", "The score turned into words.",
     "Very Strong, Strong, Neutral, Weak, Very Weak."),
    ("Trend", "Where the price sits against all four averages at once.",
     "'Above every average' is the cleanest possible setup."),
    ("Vol vs 20d Avg", "Today's volume divided by its 20-day average.",
     "2.00x means twice the usual interest. Big moves on 0.5x rarely hold."),
    ("RSI (14)", "Momentum on a 0 to 100 scale.",
     "Over 70 is stretched, under 30 is beaten down, 50 is the dividing line."),
    ("Stoch %K", "Where the close sits inside the last 14 days' range.",
     "100 means it closed at the top of the range, 0 at the very bottom."),
    ("ADX (14)", "How strong the trend is, whichever way it points.",
     "Over 25 means a real trend. Under 20 means it is drifting sideways."),
    ("MACD / Signal / Histogram", "The three raw MACD numbers off the chart.",
     "Histogram is MACD minus Signal. It crossing zero is the actual event."),
    ("%B", "Where the price sits inside the Bollinger Bands.",
     "1.00 is the upper band, 0.00 is the lower, 0.50 is the middle."),
    ("Band Width", "How far apart the bands are, as a percent of the middle.",
     "Narrow bands come before big moves. Wide bands come after them."),
    ("ATR (14)", "The average daily range, as a percent of price.",
     "A 4% ATR means 4% swings are normal for that name. It sizes the risk."),
    ("% Off 52W High", "How far below the year's high it is trading.",
     "Zero means it is at a new high right now."),
    ("52W Range Position", "Where it sits between the year's low and high.",
     "100% is at the high, 0% is at the low."),
    ("5-Day / 1-Month / 3-Month / YTD", "Returns over longer windows.",
     "One green day inside four red columns is noise, not a turn."),
    ("Top News 1-3", "The three most recent headlines. Click to open.",
     "Straight from Yahoo Finance. Useful for explaining an odd move."),
    ("", "", ""),
    ("THE OTHER TABS", "", ""),
    ("Dashboard", "The whole watchlist in one screen.",
     "Advancers, decliners, breadth, sector ranking, leaders and laggards."),
    ("Sector Rotation", "The eleven sectors across six time windows.",
     "Green all the way across is a real trend. Green in one column is a bounce."),
    ("Signals", "Only the tickers where something actually happened.",
     "Crossovers, band breaks, volume spikes, new highs and lows."),
]


def build_guide(wb: Workbook):
    ws = wb.create_sheet("How To Read This")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "475569"
    for col, w in zip("ABCD", [3, 30, 58, 66]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:D2")
    put(ws, 2, 2, "HOW TO READ THIS WORKBOOK", bold=True, size=18, colour="FFFFFF",
        fill=INK, align="left", border=False)
    for c in range(2, 5):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=INK)
    ws.row_dimensions[2].height = 30
    put(ws, 3, 2, "Every column, in plain English. Nothing here needs to be typed in "
                  "by hand any more.", size=10.5, colour=LABEL_GREY, italic=True,
        border=False)

    r = 5
    for label, what, note in GUIDE:
        if not label and not what:
            r += 1
            continue
        if not what:                                    # a section heading
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            put(ws, r, 2, label, bold=True, size=12, colour="FFFFFF",
                fill="1E3A5F", align="left", border=False)
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="1E3A5F")
            ws.row_dimensions[r].height = 22
            r += 1
            continue
        put(ws, r, 2, label, bold=True, size=10.5, colour=INK, align="left")
        put(ws, r, 3, what, size=10.5, colour="1E293B", align="left", wrap=True)
        put(ws, r, 4, note, size=10, colour=LABEL_GREY, align="left", wrap=True,
            italic=True)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    put(ws, r, 2, "Colour key", bold=True, size=11, colour=INK, border=False)
    r += 1
    for text, bg, fg, meaning in (
        ("Yes", GREEN_BG, GREEN_FG, "The answer is yes. Bullish."),
        ("No", RED_BG, RED_FG, "The answer is no. Bearish."),
        ("Same", AMBER_BG, AMBER_FG, "Too close to call — within a tenth of a percent."),
    ):
        put(ws, r, 2, text, bold=True, size=10.5, colour=fg, fill=bg, align="center")
        put(ws, r, 3, meaning, size=10.5, colour="1E293B", align="left")
        r += 1
    return ws


# ==============================================================================
#  ASSEMBLING THE WORKBOOK
# ==============================================================================

def build_rows(watchlist: Dict[str, Any], frames: Dict[str, pd.DataFrame],
               asof: pd.Timestamp, cfg: Dict[str, Any],
               profiles: Dict[str, Dict]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for group in watchlist["groups"]:
        analysed = []
        for ticker in group["tickers"]:
            df = frames.get(ticker)
            a = analyse(ticker, df, asof, cfg) if df is not None else None
            analysed.append({
                "ticker": ticker,
                "group": group["key"],
                "group_title": group["title"],
                "numbered": group.get("numbered", False),
                "name": profiles.get(ticker, {}).get("name", ticker),
                "a": a,
            })
        if group.get("rank", True):
            scored = [x for x in analysed if x["a"]]
            scored.sort(key=lambda x: x["a"]["pct_change"], reverse=True)
            for n, x in enumerate(scored, start=1):
                x["strength"] = n
        rows.extend(analysed)
    return rows


def trading_days(frames: Dict[str, pd.DataFrame], count: int) -> List[pd.Timestamp]:
    """The most recent N session dates that appear across the downloaded data."""
    index = None
    for df in frames.values():
        index = df.index if index is None else index.union(df.index)
    if index is None:
        return []
    return list(index.sort_values()[-count:][::-1])


def generate(watchlist: Dict[str, Any], out_path: str, history: int = 5,
             want_news: bool = True, json_path: Optional[str] = None) -> Dict[str, Any]:
    cfg = {**DEFAULT_WATCHLIST["settings"], **watchlist.get("settings", {})}
    tickers = [t for g in watchlist["groups"] for t in g["tickers"]]

    print("Stock Analysis — building today's workbook")
    frames = download_history(tickers)
    if not frames:
        sys.exit("No market data came back. Check the internet connection and retry.")
    profiles = download_profiles(tickers)
    news = download_news(tickers, cfg["news_per_ticker"]) if want_news else {}

    days = trading_days(frames, max(history, 1))
    generated = dt.datetime.now().strftime("%d %b %Y, %I:%M %p").lstrip("0")

    wb = Workbook()
    wb.remove(wb.active)

    print("  Writing sheets ...")
    day_rows: Dict[str, List[Dict[str, Any]]] = {}
    for n, day in enumerate(days):
        rows = build_rows(watchlist, frames, day, cfg, profiles)
        label = day.strftime("%m-%d")
        day_rows[label] = rows
        build_day_sheet(wb, label, rows,
                        day.strftime("%A, %d %B %Y"), generated, cfg,
                        news if n == 0 else {}, profiles)

    today_label = days[0].strftime("%m-%d")
    today_rows = day_rows[today_label]
    today_pretty = days[0].strftime("%A, %d %B %Y")

    build_dashboard(wb, today_rows, today_pretty, generated, cfg)
    build_rotation(wb, today_rows, today_pretty)
    build_signals(wb, today_rows, today_pretty)
    build_guide(wb)

    wb.active = 1                                  # land on today's tab
    wb.save(out_path)
    print(f"  Saved {out_path}")

    payload = to_json(today_rows, today_pretty, days[0].strftime("%Y-%m-%d"),
                      generated, cfg, news, profiles)
    if json_path:
        with open(json_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, separators=(",", ":"))
        print(f"  Saved {json_path}")
    return payload


def to_json(rows: List[Dict[str, Any]], pretty_date: str, iso_date: str,
            generated: str, cfg: Dict[str, Any],
            news: Dict[str, List[Dict[str, str]]],
            profiles: Dict[str, Dict]) -> Dict[str, Any]:
    """Everything the website needs, in one small file."""
    out_rows = []
    for row in rows:
        a = row.get("a")
        if not a:
            continue
        item = {
            "ticker": row["ticker"],
            "group": row["group"],
            "groupTitle": row["group_title"],
            "numbered": row["numbered"],
            "name": row.get("name", row["ticker"]),
            "exchange": profiles.get(row["ticker"], {}).get("exchange", ""),
            "marketCap": profiles.get(row["ticker"], {}).get("market_cap"),
            "strength": row.get("strength"),
            "news": news.get(row["ticker"], []),
        }
        item.update({k: safe(v) if isinstance(v, (int, float)) else v
                     for k, v in a.items()
                     if k not in ("events", "spark_dates")})
        item["events"] = a.get("events", [])
        out_rows.append(item)

    spark_dates: List[str] = []
    for row in rows:
        dates = (row.get("a") or {}).get("spark_dates") or []
        if len(dates) > len(spark_dates):
            spark_dates = dates

    live = [r for r in out_rows]
    up = sum(1 for r in live if (r["pct_change"] or 0) > 0)
    down = sum(1 for r in live if (r["pct_change"] or 0) < 0)
    return {
        "date": iso_date,
        "datePretty": pretty_date,
        "generated": generated,
        "generatedIso": dt.datetime.now().isoformat(timespec="seconds"),
        "settings": cfg,
        "sparkDates": spark_dates,
        "summary": {
            "count": len(live),
            "advancing": up,
            "declining": down,
            "unchanged": len(live) - up - down,
            "breadth": round(up / max(up + down, 1) * 100, 1),
            "avgScore": round(sum(r["score"] for r in live) / max(len(live), 1), 2),
            "above200": sum(1 for r in live
                            if r.get("sma200") and r["last"] > r["sma200"]),
            "signals": sum(len(r["events"]) for r in live),
        },
        "rows": out_rows,
    }


# ==============================================================================
#  ENTRY POINT
# ==============================================================================

def load_watchlist(path: Optional[str]) -> Dict[str, Any]:
    candidates = [path] if path else []
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   "watchlist.json"))
    candidates.append("watchlist.json")
    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            with open(candidate, encoding="utf-8") as fh:
                data = json.load(fh)
            print(f"  Watchlist: {candidate}")
            return data
    print("  Watchlist: built-in default")
    return DEFAULT_WATCHLIST


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build the daily stock analysis workbook.")
    parser.add_argument("--out", help="output .xlsx path")
    parser.add_argument("--json", dest="json_path", help="also write a JSON summary")
    parser.add_argument("--history", type=int, default=5,
                        help="how many recent sessions to include as tabs (default 5)")
    parser.add_argument("--no-news", action="store_true", help="skip headline fetching")
    parser.add_argument("--watchlist", help="path to a watchlist.json")
    args = parser.parse_args()

    watchlist = load_watchlist(args.watchlist)
    out = args.out or f"Stock_Analysis_{dt.date.today():%m-%d}.xlsx"

    generate(watchlist, out, history=max(1, args.history),
             want_news=not args.no_news, json_path=args.json_path)
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
